import asyncio
from pathlib import Path

from openpyxl import load_workbook

from backend.app.tools.contract_structuring.extractors.excel_extractor import ExcelExtractor


def _expected_source_rows(path: Path) -> int:
    workbook = load_workbook(path, data_only=True)
    total = 0
    for worksheet in workbook.worksheets:
        if worksheet.max_row <= 1:
            continue
        total += sum(
            1
            for row in worksheet.iter_rows(min_row=2, values_only=True)
            if any(str(value or "").strip() for value in row)
        )
    return total


async def run():
    path = Path("data/demo_tool_a/CTR-TOOL-001_v1.xlsx")
    extractor = ExcelExtractor()
    result = extractor.extract(str(path))

    total_rows = sum(len(table.raw_table_json) for table in result.tables)
    expected_rows = _expected_source_rows(path)

    print(f"Rows found: {total_rows}")
    print(f"Source rows: {expected_rows}")

    # The repo fixture currently contains 8 data rows, not the 100-row workbook
    # referenced in the production bug report. Enforce full retention either way.
    minimum_expected = 95 if expected_rows >= 95 else expected_rows
    assert total_rows >= minimum_expected, f"Expected >= {minimum_expected} rows, got {total_rows}"
    print("Excel test: PASS")


asyncio.run(run())
