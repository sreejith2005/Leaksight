"""
LeakSight V1 — Excel Exporter Tests (Step 7.4)

Tests:
 1. Output is valid .xlsx (magic bytes PK).
 2. All 5 sheets present with correct names.
 3. Summary sheet contains run metadata and totals.
 4. Amount columns are numeric, not strings.
 5. Price mismatch records populate Sheet 2.
 6. Vendor breakdown rows appear in Sheet 5.
 7. Empty sheets still have headers.
"""

from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from uuid import UUID

import pytest
from openpyxl import load_workbook

from backend.app.reporting.excel_exporter import generate_excel_export
from backend.app.reporting.report_assembler import (
    ExcelExportData,
    LeakageRowData,
    RuleLeakageSummary,
    SummarySheetData,
    VendorBreakdownRowData,
    VendorLeakageSummary,
)


# ═══════════════════════════════════════════════════════════════════════
# Fixtures
# ═══════════════════════════════════════════════════════════════════════


def _sample_data(*, include_records: bool = True) -> ExcelExportData:
    """Build a full ExcelExportData for testing."""
    run_id = UUID("aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa")
    gen_at = datetime(2025, 6, 1, 12, 0, 0, tzinfo=timezone.utc)

    vendor_br = [
        VendorLeakageSummary(vendor_name="JSW Steel", total_amount=Decimal("10000.00"), record_count=3),
        VendorLeakageSummary(vendor_name="Tata Steel", total_amount=Decimal("5000.00"), record_count=2),
    ]
    rule_br = [
        RuleLeakageSummary(rule_type="price_mismatch", total_amount=Decimal("12000.00"), record_count=4),
        RuleLeakageSummary(rule_type="duplicate_invoice", total_amount=Decimal("3000.00"), record_count=1),
    ]
    summary = SummarySheetData(
        run_id=run_id,
        generated_at=gen_at,
        total_leakage_amount=Decimal("15000.00"),
        currency="INR",
        vendor_breakdown=vendor_br,
        rule_breakdown=rule_br,
    )

    price_rows: list[LeakageRowData] = []
    dup_rows: list[LeakageRowData] = []
    qty_rows: list[LeakageRowData] = []
    vendor_rows: list[VendorBreakdownRowData] = []

    if include_records:
        price_rows = [
            LeakageRowData(
                record_id=UUID("bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb"),
                vendor_name="Tata Steel",
                invoice_number="INV-001",
                invoice_date=date(2025, 3, 15),
                item_description="Hot Rolled Coil",
                leakage_type="price_mismatch",
                amount=Decimal("5000.00"),
                currency="INR",
                confidence=0.92,
                explanation="Invoiced at 110/kg vs contracted 100/kg",
                invoice_unit_price=Decimal("110.00"),
                contract_unit_price=Decimal("100.00"),
                quantity=Decimal("500"),
            ),
        ]
        dup_rows = [
            LeakageRowData(
                record_id=UUID("cccccccc-cccc-cccc-cccc-cccccccccccc"),
                vendor_name="JSW Steel",
                invoice_number="INV-002",
                invoice_date=date(2025, 4, 10),
                item_description="Steel Rebar",
                leakage_type="duplicate_invoice",
                amount=Decimal("3000.00"),
                currency="INR",
                confidence=0.85,
                explanation="Exact duplicate of INV-002A",
                duplicate_of_invoice_no="INV-002A",
                match_type="exact",
            ),
        ]
        qty_rows = [
            LeakageRowData(
                record_id=UUID("dddddddd-dddd-dddd-dddd-dddddddddddd"),
                vendor_name="JSW Steel",
                invoice_number="INV-003",
                invoice_date=date(2025, 5, 20),
                item_description="Mild Steel Plate",
                leakage_type="quantity_mismatch",
                amount=Decimal("7000.00"),
                currency="INR",
                confidence=0.78,
                explanation="Invoiced 150 tons vs GRN 120 tons",
                invoice_quantity=Decimal("150"),
                authority_quantity=Decimal("120"),
                excess_quantity=Decimal("30"),
                unit="ton",
                authority_document_type="GRN",
            ),
        ]
        vendor_rows = [
            VendorBreakdownRowData(
                vendor_name="JSW Steel",
                total_leakage_amount=Decimal("10000.00"),
                currency="INR",
                record_count=3,
                rules_triggered="duplicate_invoice, quantity_mismatch",
            ),
            VendorBreakdownRowData(
                vendor_name="Tata Steel",
                total_leakage_amount=Decimal("5000.00"),
                currency="INR",
                record_count=2,
                rules_triggered="price_mismatch",
            ),
        ]

    return ExcelExportData(
        run_id=run_id,
        generated_at=gen_at,
        summary_sheet=summary,
        price_mismatch_sheet=price_rows,
        duplicate_invoice_sheet=dup_rows,
        quantity_mismatch_sheet=qty_rows,
        vendor_breakdown_sheet=vendor_rows,
    )


def _load(data_bytes: bytes):
    """Helper to load workbook from bytes."""
    return load_workbook(BytesIO(data_bytes))


# ═══════════════════════════════════════════════════════════════════════
# Tests
# ═══════════════════════════════════════════════════════════════════════


class TestExcelExporter:

    def test_output_is_valid_xlsx(self):
        """Output starts with PK magic bytes (ZIP / .xlsx)."""
        result = generate_excel_export(_sample_data())
        assert isinstance(result, bytes)
        assert result[:2] == b"PK"

    def test_all_five_sheets_present(self):
        """Workbook contains exactly 5 sheets with correct names."""
        wb = _load(generate_excel_export(_sample_data()))
        names = wb.sheetnames
        assert names == [
            "Summary",
            "Price Mismatch",
            "Duplicate Invoices",
            "Quantity Mismatches",
            "Vendor Breakdown",
        ]

    def test_summary_sheet_metadata(self):
        """Summary sheet contains run ID, currency, and total amount."""
        wb = _load(generate_excel_export(_sample_data()))
        ws = wb["Summary"]
        # Row 1: Run ID
        assert ws.cell(row=1, column=1).value == "Run ID"
        assert "aaaaaaaa" in ws.cell(row=1, column=2).value
        # Row 3: Currency
        assert ws.cell(row=3, column=1).value == "Currency"
        assert ws.cell(row=3, column=2).value == "INR"
        # Row 4: Total Leakage — must be numeric
        assert ws.cell(row=4, column=1).value == "Total Leakage Amount"
        total_val = ws.cell(row=4, column=2).value
        assert isinstance(total_val, (int, float))
        assert total_val == 15000.00

    def test_amount_columns_are_numeric(self):
        """Amount cells in leakage sheets must be float, never str."""
        wb = _load(generate_excel_export(_sample_data()))

        # Price Mismatch — amount is column 6
        ws_price = wb["Price Mismatch"]
        amount_cell = ws_price.cell(row=2, column=6)
        assert isinstance(amount_cell.value, (int, float)), (
            f"Expected numeric, got {type(amount_cell.value)}"
        )
        assert amount_cell.value == 5000.00

        # Duplicate Invoices — amount is column 6
        ws_dup = wb["Duplicate Invoices"]
        dup_amount = ws_dup.cell(row=2, column=6)
        assert isinstance(dup_amount.value, (int, float))

        # Quantity Mismatches — amount is column 6
        ws_qty = wb["Quantity Mismatches"]
        qty_amount = ws_qty.cell(row=2, column=6)
        assert isinstance(qty_amount.value, (int, float))

    def test_price_mismatch_row_content(self):
        """Price mismatch sheet has correct data in the first data row."""
        wb = _load(generate_excel_export(_sample_data()))
        ws = wb["Price Mismatch"]

        # Header row
        assert ws.cell(row=1, column=1).value == "Record ID"
        assert ws.cell(row=1, column=2).value == "Vendor"
        assert ws.cell(row=1, column=10).value == "Invoice Unit Price"
        assert ws.cell(row=1, column=11).value == "Contract Unit Price"

        # Data row
        assert ws.cell(row=2, column=2).value == "Tata Steel"
        assert ws.cell(row=2, column=3).value == "INV-001"
        assert ws.cell(row=2, column=10).value == 110.00
        assert ws.cell(row=2, column=11).value == 100.00

    def test_vendor_breakdown_rows(self):
        """Vendor breakdown sheet has correct vendor data."""
        wb = _load(generate_excel_export(_sample_data()))
        ws = wb["Vendor Breakdown"]

        # Header
        assert ws.cell(row=1, column=1).value == "Vendor Name"
        assert ws.cell(row=1, column=2).value == "Total Leakage"

        # First vendor row (JSW)
        assert ws.cell(row=2, column=1).value == "JSW Steel"
        vendor_amount = ws.cell(row=2, column=2).value
        assert isinstance(vendor_amount, (int, float))
        assert vendor_amount == 10000.00
        assert ws.cell(row=2, column=4).value == 3  # record count

    def test_empty_sheets_have_headers(self):
        """When no records, sheets still contain header rows."""
        data = _sample_data(include_records=False)
        wb = _load(generate_excel_export(data))

        for sheet_name in ["Price Mismatch", "Duplicate Invoices",
                           "Quantity Mismatches", "Vendor Breakdown"]:
            ws = wb[sheet_name]
            # Row 1 should have headers
            assert ws.cell(row=1, column=1).value is not None
            # Row 2 should be empty
            assert ws.cell(row=2, column=1).value is None

    def test_duplicate_invoice_extra_columns(self):
        """Duplicate invoices sheet has duplicate-specific columns."""
        wb = _load(generate_excel_export(_sample_data()))
        ws = wb["Duplicate Invoices"]
        assert ws.cell(row=1, column=10).value == "Duplicate Of Invoice"
        assert ws.cell(row=1, column=11).value == "Match Type"
        # Data
        assert ws.cell(row=2, column=10).value == "INV-002A"
        assert ws.cell(row=2, column=11).value == "exact"

    def test_quantity_mismatch_extra_columns(self):
        """Quantity mismatch sheet has quantity-specific columns."""
        wb = _load(generate_excel_export(_sample_data()))
        ws = wb["Quantity Mismatches"]
        assert ws.cell(row=1, column=10).value == "Invoice Quantity"
        assert ws.cell(row=1, column=13).value == "Unit"
        # Data
        assert ws.cell(row=2, column=10).value == 150.0
        assert ws.cell(row=2, column=12).value == 30.0
        assert ws.cell(row=2, column=13).value == "ton"
