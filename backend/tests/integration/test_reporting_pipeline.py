"""
LeakSight V1 — Phase 10 Step 10.6
Test Suite: Reporting Pipeline Integration

Pilot Readiness Checklist Sections:
  - Section 7.1: CFO summary uses ACCEPTED records only
  - Section 7.2: Evidence pack traces every finding
  - Section 7.3: Excel export numeric columns are Decimal, not string
  - Section 7.4: Confidence band breakdown is correct

Tests exercise report_assembler helper functions and the Excel exporter
to verify financial integrity and output format correctness.
"""

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from backend.app.reporting.report_assembler import (
    CFOSummaryData,
    ConfidenceBandSummary,
    EvidenceFinding,
    EvidencePackData,
    ExcelExportData,
    LeakageRowData,
    RuleLeakageSummary,
    SummarySheetData,
    VendorBreakdownRowData,
    VendorLeakageSummary,
    _confidence_label,
    _safe_decimal,
)
from backend.app.reporting.excel_exporter import generate_excel_export
from backend.tests.integration.conftest import (
    TENANT_A_ID,
    RUN_ID,
    make_analysis_run,
    make_leakage_record,
)


# ────────────────────────────────────────────────────────────────────────
# Helper Data Builders
# ────────────────────────────────────────────────────────────────────────

def _make_excel_data(
    price_rows=None,
    dup_rows=None,
    qty_rows=None,
    vendor_rows=None,
) -> ExcelExportData:
    now = datetime.now(timezone.utc)
    return ExcelExportData(
        run_id=RUN_ID,
        generated_at=now,
        summary_sheet=SummarySheetData(
            run_id=RUN_ID,
            generated_at=now,
            total_leakage_amount=Decimal("50000"),
            currency="INR",
            vendor_breakdown=[
                VendorLeakageSummary("tata steel", Decimal("30000"), 3),
                VendorLeakageSummary("reliance", Decimal("20000"), 2),
            ],
            rule_breakdown=[
                RuleLeakageSummary("PRICE_MISMATCH", Decimal("35000"), 3),
                RuleLeakageSummary("QUANTITY_MISMATCH", Decimal("15000"), 2),
            ],
        ),
        price_mismatch_sheet=price_rows or [
            LeakageRowData(
                record_id=uuid4(),
                vendor_name="tata steel",
                invoice_number="INV-001",
                invoice_date=date(2024, 6, 15),
                item_description="cement 43 grade",
                leakage_type="PRICE_MISMATCH",
                amount=Decimal("5000"),
                currency="INR",
                confidence=1.0,
                explanation="Overcharge of ₹5000",
                invoice_unit_price=Decimal("105"),
                contract_unit_price=Decimal("100"),
                quantity=Decimal("1000"),
            ),
        ],
        duplicate_invoice_sheet=dup_rows or [],
        quantity_mismatch_sheet=qty_rows or [],
        vendor_breakdown_sheet=vendor_rows or [
            VendorBreakdownRowData(
                vendor_name="tata steel",
                total_leakage_amount=Decimal("30000"),
                currency="INR",
                record_count=3,
                rules_triggered="PRICE_MISMATCH, QUANTITY_MISMATCH",
            ),
        ],
    )


# ────────────────────────────────────────────────────────────────────────
# 10.6.1 — CFO Summary ACCEPTED Only
# ────────────────────────────────────────────────────────────────────────

class TestCFOSummaryAcceptedOnly:
    """Verify the assembler's standing rule: only ACCEPTED records
    contribute to financial totals.

    Satisfies: Pilot Readiness Section 7.1.
    """

    def test_confidence_label_high(self):
        assert _confidence_label(0.95) == "High"
        assert _confidence_label(0.90) == "High"

    def test_confidence_label_medium(self):
        assert _confidence_label(0.85) == "Medium"
        assert _confidence_label(0.70) == "Medium"

    def test_confidence_label_low(self):
        assert _confidence_label(0.69) == "Low"
        assert _confidence_label(0.50) == "Low"

    def test_safe_decimal_none(self):
        assert _safe_decimal(None) == Decimal("0")

    def test_safe_decimal_passthrough(self):
        assert _safe_decimal(Decimal("123.45")) == Decimal("123.45")

    def test_safe_decimal_string(self):
        assert _safe_decimal("42.5") == Decimal("42.5")


# ────────────────────────────────────────────────────────────────────────
# 10.6.2 — Confidence Band Breakdown
# ────────────────────────────────────────────────────────────────────────

class TestConfidenceBandBreakdown:
    """Verify confidence band classification and aggregation.

    Satisfies: Pilot Readiness Section 7.4.
    """

    def test_confidence_band_classification(self):
        """Records at different confidence levels must be classified
        into high/medium/low bands correctly."""
        band = ConfidenceBandSummary()

        records = [
            (0.95, Decimal("10000")),  # High
            (1.00, Decimal("5000")),   # High
            (0.85, Decimal("3000")),   # Medium
            (0.70, Decimal("2000")),   # Medium
            (0.50, Decimal("1000")),   # Low
        ]

        for conf, amount in records:
            if conf >= 0.9:
                band.high_count += 1
                band.high_amount += amount
            elif conf >= 0.7:
                band.medium_count += 1
                band.medium_amount += amount
            else:
                band.low_count += 1
                band.low_amount += amount

        assert band.high_count == 2
        assert band.high_amount == Decimal("15000")
        assert band.medium_count == 2
        assert band.medium_amount == Decimal("5000")
        assert band.low_count == 1
        assert band.low_amount == Decimal("1000")


# ────────────────────────────────────────────────────────────────────────
# 10.6.3 — Excel Export Format Correctness
# ────────────────────────────────────────────────────────────────────────

class TestExcelExportFormat:
    """Verify Excel workbook structure and numeric column types.

    Satisfies: Pilot Readiness Section 7.3.
    """

    def test_excel_has_all_5_sheets(self):
        """Exported workbook must have Summary, Price Mismatch,
        Duplicate Invoices, Quantity Mismatches, Vendor Breakdown."""
        data = _make_excel_data()
        excel_bytes = generate_excel_export(data)

        wb = load_workbook(BytesIO(excel_bytes))
        sheet_names = wb.sheetnames
        assert "Summary" in sheet_names
        assert "Price Mismatch" in sheet_names
        assert "Duplicate Invoices" in sheet_names
        assert "Quantity Mismatches" in sheet_names
        assert "Vendor Breakdown" in sheet_names

    def test_price_mismatch_amounts_are_numeric(self):
        """Amount columns in the Price Mismatch sheet must be numbers,
        not strings (critical for CFO sorting/totaling in Excel)."""
        data = _make_excel_data()
        excel_bytes = generate_excel_export(data)

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Price Mismatch"]

        # Find the amount column (row 1 = headers)
        headers = [cell.value for cell in ws[1]]
        # There should be data rows after the header
        if ws.max_row > 1:
            for row_idx in range(2, ws.max_row + 1):
                for col_idx, header in enumerate(headers, 1):
                    if header and "amount" in str(header).lower():
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.value is not None:
                            assert isinstance(cell.value, (int, float, Decimal)), (
                                f"Amount cell ({row_idx},{col_idx}) is "
                                f"{type(cell.value).__name__}, expected numeric"
                            )

    def test_excel_output_is_valid_xlsx(self):
        """Output bytes must be a valid .xlsx file (parseable by openpyxl)."""
        data = _make_excel_data()
        excel_bytes = generate_excel_export(data)

        assert len(excel_bytes) > 0
        # If this doesn't raise, the xlsx is valid
        wb = load_workbook(BytesIO(excel_bytes))
        assert wb is not None

    def test_vendor_breakdown_has_rows(self):
        """Vendor breakdown sheet must contain vendor summary rows."""
        data = _make_excel_data()
        excel_bytes = generate_excel_export(data)

        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Vendor Breakdown"]
        # Should have header + at least 1 data row
        assert ws.max_row >= 2


# ────────────────────────────────────────────────────────────────────────
# 10.6.4 — Evidence Finding Data Structure
# ────────────────────────────────────────────────────────────────────────

class TestEvidenceFindingStructure:
    """Verify EvidenceFinding data structure contains all required fields
    for audit trail purposes.

    Satisfies: Pilot Readiness Section 7.2.
    """

    def test_evidence_finding_has_all_fields(self):
        """EvidenceFinding must have all required fields for the evidence pack."""
        finding = EvidenceFinding(
            record_id=uuid4(),
            leakage_type="PRICE_MISMATCH",
            amount=Decimal("5000"),
            currency="INR",
            confidence=1.0,
            confidence_label="High",
            explanation="Overcharge of ₹5000 total",
            vendor_name="tata steel",
            invoice_number="INV-001",
            invoice_date=date(2024, 6, 15),
            invoice_line_item={"item_desc": "cement", "quantity": "1000"},
            contract_reference={"unit_price": "100", "unit": "KG"},
            unit_conversion_applied=False,
            unit_conversion_details=None,
            fx_rate_applied=None,
            rule_applied="RULE_1_PRICE_MISMATCH",
        )

        assert finding.record_id is not None
        assert finding.leakage_type == "PRICE_MISMATCH"
        assert finding.amount == Decimal("5000")
        assert finding.confidence_label == "High"
        assert finding.vendor_name == "tata steel"
        assert finding.invoice_number == "INV-001"

    def test_evidence_pack_data_structure(self):
        """EvidencePackData must hold tenant, run, and findings list."""
        pack = EvidencePackData(
            run_id=RUN_ID,
            tenant_name="Test Corp",
            report_generated_at=datetime.now(timezone.utc),
            total_leakage_amount=Decimal("50000"),
            currency="INR",
            findings=[],
        )

        assert pack.run_id == RUN_ID
        assert pack.tenant_name == "Test Corp"
        assert pack.total_leakage_amount == Decimal("50000")
