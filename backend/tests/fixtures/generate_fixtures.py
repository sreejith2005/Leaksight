"""
Generate test fixture files for Excel/CSV parser tests.

Run: python -m backend.tests.fixtures.generate_fixtures
"""

import csv
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

FIXTURE_DIR = Path(__file__).parent


def create_clean_excel():
    """Clean XLSX with header on row 1 and 5 line items."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    headers = ["Item Description", "Quantity", "Unit", "Unit Price", "Amount"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    data = [
        ("Cement OPC 53 Grade", 100, "Bags", 350.00, 35000.00),
        ("TMT Steel 12mm", 500, "Kg", 72.50, 36250.00),
        ("River Sand", 10, "Cum", 2800.00, 28000.00),
        ("Bricks First Class", 5000, "Nos", 8.50, 42500.00),
        ("Plywood 18mm", 20, "Sheets", 1450.00, 29000.00),
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(FIXTURE_DIR / "clean_invoice.xlsx")


def create_clean_csv():
    """Clean CSV with header on row 1 and 3 line items."""
    rows = [
        ["Description", "Qty", "UOM", "Rate", "Total"],
        ["Cement PPC", 200, "Bags", 320.00, 64000.00],
        ["Sand Fine", 15, "Cum", 2500.00, 37500.00],
        ["Aggregate 20mm", 25, "Cum", 1800.00, 45000.00],
    ]
    with open(FIXTURE_DIR / "clean_invoice.csv", "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def create_header_on_row3():
    """Excel with vendor info on rows 1-2 and header on row 3."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value="Vendor: ABC Suppliers Pvt Ltd")
    ws.cell(row=2, column=1, value="Invoice No: INV-2024-0451")
    ws.cell(row=2, column=3, value="Date: 15/03/2024")

    headers = ["Item", "Quantity", "Unit", "Price", "Amount"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    data = [
        ("20mm Aggregate", 30, "Cum", 1700.00, 51000.00),
        ("M-Sand", 20, "Cum", 2200.00, 44000.00),
    ]
    for row_idx, row_data in enumerate(data, 4):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(FIXTURE_DIR / "header_row3.xlsx")


def create_merged_cells():
    """Excel with merged header cells (e.g. title spanning columns)."""
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.merge_cells("A1:E1")
    ws.cell(row=1, column=1, value="Purchase Order - Material Supply")

    ws.merge_cells("A2:C2")
    ws.cell(row=2, column=1, value="Vendor: XYZ Materials")

    headers = ["Description", "Qty", "Unit", "Rate", "Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    data = [
        ("Portland Cement", 50, "Bags", "₹380.00", "₹19,000.00"),
        ("Fly Ash", 100, "Kg", "₹12.50", "₹1,250.00"),
    ]
    for row_idx, row_data in enumerate(data, 4):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(FIXTURE_DIR / "merged_cells.xlsx")


def create_malformed():
    """Completely malformed Excel - binary noise."""
    with open(FIXTURE_DIR / "malformed.xlsx", "wb") as f:
        f.write(b"\x00\x01\x02\x03NOTANEXCEL\xff\xfe\xfd")


if __name__ == "__main__":
    create_clean_excel()
    create_clean_csv()
    create_header_on_row3()
    create_merged_cells()
    create_malformed()
    print("Fixtures generated successfully.")
