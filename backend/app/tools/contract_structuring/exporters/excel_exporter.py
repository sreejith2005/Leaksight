"""Structured Excel export for Tool A."""

from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADERS = [
    "Vendor Name",
    "Contract ID",
    "Effective Date",
    "Expiry Date",
    "Item Description",
    "Unit",
    "Unit Price",
    "Currency",
    "Version",
    "Source Page",
    "Confidence",
    "Needs Review",
]


def _format_date(value: date | None) -> str:
    if value is None:
        return ""
    return value.strftime("%d/%m/%Y")


def export_structuring_excel(
    output_path: str | Path,
    document_rows: list[dict[str, Any]],
) -> str:
    """Create Tool A structured Excel output with one sheet per document."""
    wb = Workbook()
    wb.remove(wb.active)

    for block in document_rows:
        sheet_name = str(block.get("sheet_name", "Contract"))[:31]
        ws = wb.create_sheet(sheet_name)
        ws.append(_HEADERS)

        for row in block.get("rows", []):
            ws.append(
                [
                    row.get("vendor_name"),
                    row.get("contract_id"),
                    _format_date(row.get("effective_date")),
                    _format_date(row.get("expiry_date")),
                    row.get("item_description"),
                    row.get("unit_raw"),
                    float(row.get("unit_price")) if row.get("unit_price") is not None else None,
                    row.get("currency") or "INR",
                    row.get("version_number") or 1,
                    row.get("source_page"),
                    round(float(row.get("confidence") or 0.0), 2),
                    "YES" if row.get("needs_review") else "NO",
                ]
            )

        header_fill = PatternFill(fill_type="solid", start_color="C9A84C", end_color="C9A84C")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill

        ws.freeze_panes = "A2"

        for col_idx in range(1, len(_HEADERS) + 1):
            max_len = len(_HEADERS[col_idx - 1])
            for row_idx in range(2, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    output_path = str(output_path)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
