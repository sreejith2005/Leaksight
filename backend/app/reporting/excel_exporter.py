"""
LeakSight V1 — Excel Exporter

Source: docs/ARCHITECTURE.md (Section 6.7),
       docs/DECISIONS.md — openpyxl only, never pandas for output,
       docs/API_CONTRACTS.md (Section 7 — export-excel endpoint)

Generates an Excel workbook from ExcelExportData assembled by the report
assembler.  Output is returned as ``bytes`` from a ``BytesIO`` buffer —
**never writes to disk**.

Amount columns must be numeric in the workbook (never strings).
"""

from __future__ import annotations

import logging
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.app.reporting.report_assembler import (
    ExcelExportData,
    LeakageRowData,
    VendorBreakdownRowData,
)

logger = logging.getLogger("leaksight.reporting.excel_exporter")

# ─── Style constants ────────────────────────────────────────────────────
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="D0D7DE"),
)
_NUMBER_FORMAT = '#,##0.00'
_PERCENT_FORMAT = '0%'


def generate_excel_export(data: ExcelExportData) -> bytes:
    """Generate a multi-sheet Excel workbook and return raw bytes.

    Sheets:
      1. Summary — run metadata, vendor & rule breakdowns
      2. Price Mismatch — rule-1 leakage records
      3. Duplicate Invoices — rule-2 leakage records
      4. Quantity Mismatches — rule-3 leakage records
      5. Vendor Breakdown — per-vendor totals

    Parameters
    ----------
    data:
        ``ExcelExportData`` from the report assembler.

    Returns
    -------
    bytes
        Raw ``.xlsx`` content suitable for streaming response.
    """
    wb = Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"  # type: ignore[union-attr]
    _write_summary_sheet(ws_summary, data)  # type: ignore[arg-type]

    # ── Sheet 2: Price Mismatch ─────────────────────────────────────
    ws_price = wb.create_sheet("Price Mismatch")
    _write_price_mismatch_sheet(ws_price, data.price_mismatch_sheet)

    # ── Sheet 3: Duplicate Invoices ─────────────────────────────────
    ws_dup = wb.create_sheet("Duplicate Invoices")
    _write_duplicate_invoice_sheet(ws_dup, data.duplicate_invoice_sheet)

    # ── Sheet 4: Quantity Mismatches ────────────────────────────────
    ws_qty = wb.create_sheet("Quantity Mismatches")
    _write_quantity_mismatch_sheet(ws_qty, data.quantity_mismatch_sheet)

    # ── Sheet 5: Vendor Breakdown ───────────────────────────────────
    ws_vendor = wb.create_sheet("Vendor Breakdown")
    _write_vendor_breakdown_sheet(ws_vendor, data.vendor_breakdown_sheet)

    # Serialize to bytes — never write to disk
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════════
# Sheet writers
# ═══════════════════════════════════════════════════════════════════════


def _style_header_row(ws, col_count: int) -> None:
    """Apply header style to the first row."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT


def _auto_width(ws, col_count: int) -> None:
    """Auto-fit column widths based on header text length (simple heuristic)."""
    for col_idx in range(1, col_count + 1):
        header_val = ws.cell(row=1, column=col_idx).value or ""
        width = max(len(str(header_val)) + 4, 12)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_summary_sheet(ws, data: ExcelExportData) -> None:
    """Write summary metadata, vendor breakdown, and rule breakdown."""
    # Section 1: Run metadata
    ws.append(["Run ID", str(data.run_id)])
    ws.append(["Generated At", str(data.generated_at)])
    ws.append(["Currency", data.summary_sheet.currency])
    ws.append(["Total Leakage Amount", float(data.summary_sheet.total_leakage_amount)])
    ws.cell(row=4, column=2).number_format = _NUMBER_FORMAT
    ws.append([])

    # Section 2: Vendor Breakdown
    ws.append(["Vendor Breakdown"])
    row_start = ws.max_row + 1
    ws.append(["Vendor Name", "Leakage Amount", "Record Count"])
    _style_header_row_at(ws, row_start, 3)
    for v in data.summary_sheet.vendor_breakdown:
        ws.append([v.vendor_name, float(v.total_amount), v.record_count])
        ws.cell(row=ws.max_row, column=2).number_format = _NUMBER_FORMAT
    ws.append([])

    # Section 3: Rule Breakdown
    ws.append(["Rule Breakdown"])
    row_start = ws.max_row + 1
    ws.append(["Rule Type", "Leakage Amount", "Record Count"])
    _style_header_row_at(ws, row_start, 3)
    for r in data.summary_sheet.rule_breakdown:
        ws.append([r.rule_type, float(r.total_amount), r.record_count])
        ws.cell(row=ws.max_row, column=2).number_format = _NUMBER_FORMAT

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 15


def _style_header_row_at(ws, row: int, col_count: int) -> None:
    """Apply header style to a specific row (sub-tables in summary)."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT


def _write_leakage_base_headers() -> list[str]:
    """Common columns for all leakage sheets."""
    return [
        "Record ID",
        "Vendor",
        "Invoice Number",
        "Invoice Date",
        "Item Description",
        "Amount",
        "Currency",
        "Confidence",
        "Explanation",
    ]


def _write_leakage_base_row(row_data: LeakageRowData) -> list:
    """Common column values for all leakage sheets."""
    return [
        str(row_data.record_id),
        row_data.vendor_name,
        row_data.invoice_number,
        str(row_data.invoice_date) if row_data.invoice_date else "",
        row_data.item_description,
        float(row_data.amount),     # numeric, never string
        row_data.currency,
        row_data.confidence,        # numeric float
        row_data.explanation,
    ]


def _write_price_mismatch_sheet(ws, rows: List[LeakageRowData]) -> None:
    """Sheet for Rule-1 price mismatch records."""
    headers = _write_leakage_base_headers() + [
        "Invoice Unit Price",
        "Contract Unit Price",
        "Quantity",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for r in rows:
        base = _write_leakage_base_row(r)
        base.extend([
            float(r.invoice_unit_price) if r.invoice_unit_price is not None else "",
            float(r.contract_unit_price) if r.contract_unit_price is not None else "",
            float(r.quantity) if r.quantity is not None else "",
        ])
        ws.append(base)
        # Amount column (6) must be numeric
        ws.cell(row=ws.max_row, column=6).number_format = _NUMBER_FORMAT
        # Unit price columns
        for col in (10, 11):
            ws.cell(row=ws.max_row, column=col).number_format = _NUMBER_FORMAT

    _auto_width(ws, len(headers))


def _write_duplicate_invoice_sheet(ws, rows: List[LeakageRowData]) -> None:
    """Sheet for Rule-2 duplicate invoice records."""
    headers = _write_leakage_base_headers() + [
        "Duplicate Of Invoice",
        "Match Type",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for r in rows:
        base = _write_leakage_base_row(r)
        base.extend([
            r.duplicate_of_invoice_no or "",
            r.match_type or "",
        ])
        ws.append(base)
        ws.cell(row=ws.max_row, column=6).number_format = _NUMBER_FORMAT

    _auto_width(ws, len(headers))


def _write_quantity_mismatch_sheet(ws, rows: List[LeakageRowData]) -> None:
    """Sheet for Rule-3 quantity mismatch records."""
    headers = _write_leakage_base_headers() + [
        "Invoice Quantity",
        "Authority Quantity",
        "Excess Quantity",
        "Unit",
        "Authority Document",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for r in rows:
        base = _write_leakage_base_row(r)
        base.extend([
            float(r.invoice_quantity) if r.invoice_quantity is not None else "",
            float(r.authority_quantity) if r.authority_quantity is not None else "",
            float(r.excess_quantity) if r.excess_quantity is not None else "",
            r.unit or "",
            r.authority_document_type or "",
        ])
        ws.append(base)
        ws.cell(row=ws.max_row, column=6).number_format = _NUMBER_FORMAT

    _auto_width(ws, len(headers))


def _write_vendor_breakdown_sheet(ws, rows: List[VendorBreakdownRowData]) -> None:
    """Vendor breakdown totals sheet."""
    headers = ["Vendor Name", "Total Leakage", "Currency", "Record Count", "Rules Triggered"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for r in rows:
        ws.append([
            r.vendor_name,
            float(r.total_leakage_amount),  # numeric, never string
            r.currency,
            r.record_count,
            r.rules_triggered,
        ])
        ws.cell(row=ws.max_row, column=2).number_format = _NUMBER_FORMAT

    _auto_width(ws, len(headers))
